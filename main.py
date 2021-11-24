import pandas as pd
from openpyxl import load_workbook

from utils import folder_input, cre_shape, Change_method
from table import bor_row, min_border, split_format, drop_columns, split_table
from title_unit import split_element, print_element




def main():
    path_folder = '2020_bariavungtau - Copy'
    writer = pd.ExcelWriter('2020_bariavungtau.xlsx') #write file table output
    for i in folder_input(path_folder): #read file input
        path_excel = path_folder+"/"+i
        try:
            item = load_workbook(path_excel,data_only=True)
        except ValueError:
            item = load_workbook(path_excel)
        sheet_item = item.sheetnames
        for name in sheet_item:
            print(name)
            sheet = item[name]
            data = pd.read_excel(path_excel,sheet_name=name,na_values=[' ',','],header=None)
            if (sheet.sheet_state != 'hidden') & (data.shape[1] != 0):
                count_shape = 0
                lst_shape = cre_shape(data,path_excel,name)
                ######################
                for shape_data in lst_shape:
                    lst_border = bor_row(sheet,shape_data) #Tạo list chứa vị trí cắt.
                    #######################
                    if Change_method(len(lst_border)) == 0:
                        item_test = pd.read_excel(path_excel,sheet_name=name,usecols=range(shape_data[1][0]-1,shape_data[1][1]),na_values=[' ',','],header=None)
                        lst_border = min_border(lst_border,item_test)
                        lst_border = split_format(item_test,lst_border)
                        lst_border, item_test= drop_columns(lst_border, item_test)
                        #######################################
                        ###################################
                        item_test = split_table(item_test,count_shape,lst_border,name,writer)
                        if Change_method(item_test.shape[1]) == 0:
                            ###########Xử lý Title, Đơn vị 
                            lst_element, item_test = split_element(item_test)
                            ############## Kiểm tra Bảng xử lý vị lỗi
                            print_element(lst_element,item_test,name)
    writer.save()

if __name__ == '__main__':
    main()